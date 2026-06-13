import csv
import io
from datetime import datetime
from typing import List
from app.models.invoice import Invoice


def export_csv(invoices: List[Invoice]) -> bytes:
    """
    Generates a CSV file from a list of invoices.
    Returns bytes so FastAPI can stream it directly.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "Invoice Number", "Vendor", "Amount", "Tax Amount",
        "Currency", "Category", "Invoice Date", "Status",
        "Fraud Score", "Is Duplicate", "Source", "Uploaded On"
    ])

    for inv in invoices:
        writer.writerow([
            inv.invoice_number or "",
            inv.vendor_name or "",
            inv.total_amount or "",
            inv.tax_amount or "",
            inv.currency or "INR",
            inv.category or "",
            inv.invoice_date or "",
            inv.status or "",
            inv.fraud_score or 0.0,
            "Yes" if inv.is_duplicate else "No",
            inv.source or "",
            inv.created_at.strftime("%d %b %Y") if inv.created_at else "",
        ])

    return output.getvalue().encode("utf-8")


def export_excel(invoices: List[Invoice]) -> bytes:
    """
    Generates a formatted Excel file.
    Includes header styling, column widths, and a summary row.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # --- Styling constants ---
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    alt_fill    = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="E2E8F0")
    thin_border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )

    headers = [
        "Invoice #", "Vendor", "Amount (₹)", "Tax (₹)",
        "Currency", "Category", "Invoice Date", "Status",
        "Fraud Score", "Duplicate", "Source", "Uploaded On"
    ]
    col_widths = [14, 28, 14, 12, 10, 14, 14, 12, 12, 10, 10, 14]

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Write data rows
    for row_idx, inv in enumerate(invoices, start=2):
        row_data = [
            inv.invoice_number or "—",
            inv.vendor_name or "—",
            inv.total_amount or 0,
            inv.tax_amount or 0,
            inv.currency or "INR",
            inv.category or "—",
            str(inv.invoice_date) if inv.invoice_date else "—",
            inv.status or "—",
            inv.fraud_score or 0.0,
            "Yes" if inv.is_duplicate else "No",
            inv.source or "—",
            inv.created_at.strftime("%d %b %Y") if inv.created_at else "—",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Alternate row shading for readability
            if row_idx % 2 == 0:
                cell.fill = alt_fill

            # Highlight flagged rows in red
            if inv.is_flagged:
                cell.fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

        ws.row_dimensions[row_idx].height = 18

    # Summary row at the bottom
    summary_row = len(invoices) + 3
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    total_amount = sum(inv.total_amount or 0 for inv in invoices)
    total_tax = sum(inv.tax_amount or 0 for inv in invoices)
    ws.cell(row=summary_row, column=3, value=round(total_amount, 2)).font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=round(total_tax, 2)).font = Font(bold=True)
    ws.cell(row=summary_row, column=1, value=f"TOTAL  ({len(invoices)} invoices)").font = Font(bold=True)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def export_pdf(invoices: List[Invoice], user_name: str = "") -> bytes:
    """
    Generates a clean PDF summary report.
    Includes a header, summary stats, and a data table.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    brand_green = colors.HexColor("#16A34A")
    light_green = colors.HexColor("#F0FDF4")
    light_gray  = colors.HexColor("#F8FAFC")
    mid_gray    = colors.HexColor("#64748B")
    red_light   = colors.HexColor("#FEF2F2")

    title_style = ParagraphStyle(
        "title", fontSize=20, textColor=brand_green,
        fontName="Helvetica-Bold", spaceAfter=2
    )
    subtitle_style = ParagraphStyle(
        "subtitle", fontSize=10, textColor=mid_gray,
        fontName="Helvetica", spaceAfter=12
    )
    section_style = ParagraphStyle(
        "section", fontSize=12, textColor=colors.HexColor("#1E293B"),
        fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=6
    )

    story = []

    # --- Title block ---
    story.append(Paragraph("MailToMint", title_style))
    generated_on = datetime.utcnow().strftime("%d %B %Y, %H:%M UTC")
    story.append(Paragraph(f"Invoice Export Report · {generated_on}", subtitle_style))
    if user_name:
        story.append(Paragraph(f"Prepared for: {user_name}", subtitle_style))
    story.append(Spacer(1, 6*mm))

    # --- Summary stats table ---
    total_amount = sum(inv.total_amount or 0 for inv in invoices)
    total_tax    = sum(inv.tax_amount or 0 for inv in invoices)
    flagged      = sum(1 for inv in invoices if inv.is_flagged)
    processed    = sum(1 for inv in invoices if inv.status == "processed")

    story.append(Paragraph("Summary", section_style))
    summary_data = [
        ["Total Invoices", "Total Spend", "Total Tax", "Flagged", "Processed"],
        [
            str(len(invoices)),
            f"Rs. {total_amount:,.2f}",
            f"Rs. {total_tax:,.2f}",
            str(flagged),
            str(processed),
        ]
    ]
    summary_table = Table(summary_data, colWidths=[36*mm]*5)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), brand_green),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 9),
        ("FONTNAME",    (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,1), (-1,1), 12),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWHEIGHT",   (0,0), (-1,-1), 10*mm),
        ("BACKGROUND",  (0,1), (-1,1), light_green),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#D1FAE5")),
        ("ROUNDEDCORNERS", [3]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 6*mm))

    # --- Invoice detail table ---
    story.append(Paragraph("Invoice Details", section_style))

    table_headers = ["#", "Vendor", "Invoice No.", "Amount", "Category", "Date", "Status"]
    table_data = [table_headers]

    for i, inv in enumerate(invoices, start=1):
        table_data.append([
            str(i),
            (inv.vendor_name or "—")[:25],
            inv.invoice_number or "—",
            f"Rs.{inv.total_amount:,.0f}" if inv.total_amount else "—",
            inv.category or "—",
            str(inv.invoice_date) if inv.invoice_date else inv.created_at.strftime("%d/%m/%y"),
            inv.status or "—",
        ])

    col_widths_pdf = [10*mm, 50*mm, 30*mm, 25*mm, 25*mm, 22*mm, 20*mm]
    detail_table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)

    # Build row styles — highlight flagged rows
    table_style_cmds = [
        ("BACKGROUND",  (0,0), (-1,0), brand_green),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 8),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWHEIGHT",   (0,0), (-1,-1), 8*mm),
        ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, light_gray]),
    ]

    # Highlight flagged invoices in red
    for i, inv in enumerate(invoices, start=1):
        if inv.is_flagged:
            table_style_cmds.append(("BACKGROUND", (0,i), (-1,i), red_light))

    detail_table.setStyle(TableStyle(table_style_cmds))
    story.append(detail_table)

    # Footer note
    story.append(Spacer(1, 8*mm))
    footer_style = ParagraphStyle(
        "footer", fontSize=8, textColor=mid_gray,
        fontName="Helvetica", alignment=TA_CENTER
    )
    story.append(Paragraph(
        "Generated by MailToMint · AI-powered invoice intelligence",
        footer_style
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()